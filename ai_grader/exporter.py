"""
exporter.py — Write grading results to Excel and/or CSV.
"""

import csv
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_FLAG_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def write_results(
    results: list[dict],
    output_dir: str,
    questions: list[str],
    formats: list[str],
) -> list[str]:
    """
    Write results to the requested formats.
    Returns list of written file paths.
    """
    os.makedirs(output_dir, exist_ok=True)
    written = []

    if "excel" in formats:
        path = os.path.join(output_dir, "marks.xlsx")
        _write_excel(results, path, questions)
        written.append(path)

    if "csv" in formats:
        path = os.path.join(output_dir, "marks.csv")
        _write_csv(results, path, questions)
        written.append(path)

    return written


# ── Excel ─────────────────────────────────────────────────────────────────────

def _write_excel(results: list[dict], path: str, questions: list[str]) -> None:
    wb = openpyxl.Workbook()

    # Marks sheet
    ws = wb.active
    ws.title = "Marks"
    headers = ["student_id", "name"] + questions + ["total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        q_vals = [r.get(q, -1) for q in questions]
        total = sum(q_vals) if all(v >= 0 for v in q_vals) else -1
        row = [r["student_id"], r["name"]] + q_vals + [total]
        ws.append(row)
        if total < 0:
            for cell in ws[ws.max_row]:
                cell.fill = _FLAG_FILL

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    # Reasoning sheet
    ws2 = wb.create_sheet("Reasoning")
    reason_headers = ["student_id", "name"] + [f"{q}_reason" for q in questions]
    ws2.append(reason_headers)
    for cell in ws2[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for r in results:
        reasoning = r.get("reasoning", {})
        ws2.append(
            [r["student_id"], r["name"]]
            + [reasoning.get(q, "") for q in questions]
        )
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 40

    wb.save(path)


# ── CSV ───────────────────────────────────────────────────────────────────────

def _write_csv(results: list[dict], path: str, questions: list[str]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["student_id", "name"] + questions + ["total"])
        for r in results:
            q_vals = [r.get(q, -1) for q in questions]
            total = sum(q_vals) if all(v >= 0 for v in q_vals) else -1
            writer.writerow([r["student_id"], r["name"]] + q_vals + [total])
